import streamlit as st 
import pandas as pd
from database import Database
from auth import init_auth, check_auth, check_role, check_student_access, can_edit_student_info, check_page_access
from utils import show_success, show_error, apply_theme
from datetime import datetime, timedelta
from translations import get_text

def render():
    # Initialize authentication first
    init_auth()
    
    # Apply theme from session state
    # apply_theme()  # Disabled - using Streamlit defaults

    # Set current page for role checking
    st.session_state.current_page = "04_Lớp_học"

    check_auth()
    
    # Check if user has access to this page dynamically
    if not check_page_access('04_Lớp_học'):
        st.error("Bạn không có quyền truy cập trang này.")
        st.stop()
        return

    st.title(get_text("pages.classes.title", "Thông Tin Lớp Học"))

    db = Database()

    # Get current user's role and access level
    user_role = st.session_state.user.role
    family_student_id = st.session_state.user.family_student_id if user_role == 'family' else None
    
    # Unassigned Students Section for admin and teachers
    if user_role in ['admin', 'teacher']:
        with st.expander("👥 Học sinh chưa có lớp", expanded=False):
            unassigned_students = db.get_unassigned_students()
            
            if unassigned_students:
                st.warning(f"⚠️ Có {len(unassigned_students)} học sinh chưa được phân lớp:")
                
                for student in unassigned_students:
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        st.write(f"👨‍🎓 {student.full_name} (ID: {student.id})")
                    with col2:
                        # Quick assign to class
                        classes = db.get_classes()
                        if classes:
                            class_options = [f"{c.name} ({c.academic_year})" for c in classes]
                            selected_class = st.selectbox(
                                "Chọn lớp:",
                                options=[""] + class_options,
                                key=f"assign_class_{student.id}",
                                label_visibility="collapsed"
                            )
                    with col3:
                        if st.button("➕ Thêm vào lớp", key=f"quick_assign_{student.id}"):
                            if selected_class:
                                class_name = selected_class.split(" (")[0]
                                selected_class_obj = next((c for c in classes if c.name == class_name), None)
                                if selected_class_obj:
                                    if db.add_student_to_class(student.id, selected_class_obj.id, "Phân lớp từ danh sách chưa phân lớp"):
                                        show_success(f"Đã thêm {student.full_name} vào lớp {selected_class}")
                                        st.rerun()
                                    else:
                                        show_error("Không thể thêm học sinh vào lớp")
                            else:
                                show_error("Vui lòng chọn lớp trước khi thêm")
            else:
                st.success("✅ Tất cả học sinh đã được phân lớp")

    # Add new class button (only for admin and teachers)
    if user_role in ['admin', 'teacher'] and st.button(get_text("pages.classes.add_new_class", "➕ Thêm Lớp Mới")):
        st.session_state.adding_new_class = True

    # Form to add new class
    if user_role in ['admin', 'teacher'] and st.session_state.get('adding_new_class', False):
        with st.form("add_class_form"):
            st.subheader(get_text("pages.classes.add_class_title", "Thêm Lớp Mới"))

            # Get list of teachers
            teachers = db.get_teachers()
            teacher_options = [f"{t.full_name} (ID: {t.id})" for t in teachers]

            class_name = st.text_input(get_text("pages.classes.class_name", "Tên lớp"), key=f"new_class_name")
            teacher = st.selectbox(
                get_text("pages.classes.teacher", "Giáo viên chủ nhiệm"),
                options=teacher_options,
                format_func=lambda x: x.split(" (ID: ")[0],
                key=f"new_class_teacher"
            )
            # Academic year dropdown from 2000 to 2040
            current_year = datetime.now().year
            year_options = [f"{year}-{year+1}" for year in range(2000, 2041)]
            default_year_index = year_options.index(f"{current_year}-{current_year+1}") if f"{current_year}-{current_year+1}" in year_options else 0
            academic_year = st.selectbox(
                get_text("pages.classes.academic_year", "Năm học"), 
                options=year_options,
                index=default_year_index,
                key=f"new_class_year"
            )
            notes = st.text_area(get_text("pages.classes.notes", "Ghi chú"), key=f"new_class_notes")

            if st.form_submit_button(get_text("pages.classes.add_class_button", "Thêm lớp")):
                try:
                    teacher_id = int(teacher.split("ID: ")[1].rstrip(")"))
                    class_data = {
                        "name": class_name,
                        "teacher_id": teacher_id,
                        "academic_year": academic_year,
                        "notes": notes
                    }
                    if db.add_class(class_data):
                        show_success(get_text("pages.classes.add_success", "Thêm lớp mới thành công!"))
                        st.session_state.adding_new_class = False
                        st.rerun()
                    else:
                        show_error(get_text("pages.classes.add_error", "Không thể thêm lớp mới"))
                except Exception as e:
                    show_error(f"Lỗi: {str(e)}")

    if user_role == 'family':
        # For family users, only show their student's class
        student = db.get_students(user_role=user_role, family_student_id=family_student_id)[0]
        class_id = student.class_id
        classes = [db.get_class(class_id)] if class_id else []
    else:
        # For other users, show all classes
        classes = db.get_classes()

    if not classes:
        st.warning(get_text("pages.classes.no_class_info", "Không có thông tin lớp học."))
        return
    
    # Advanced search section - made collapsible
    with st.expander("🔍 Tìm kiếm nâng cao", expanded=False):
        with st.form("class_search_form"):
            search_col1, search_col2 = st.columns(2)
            
            with search_col1:
                class_name_search = st.text_input("Tìm theo tên lớp", key="class_name_search")
                teacher_search = st.text_input("Tìm theo tên giáo viên", key="teacher_search")
                
            with search_col2:
                academic_year_search = st.text_input("Tìm theo năm học", key="academic_year_search")
                min_students = st.number_input("Số học sinh tối thiểu", min_value=0, value=0, key="min_students")
            
            search_submitted = st.form_submit_button("🔍 Tìm kiếm lớp học", type="primary")
            
            if search_submitted:
                # Filter classes based on search criteria
                filtered_classes = []
                
                for class_obj in classes:
                    # Get teacher info
                    teacher = db.get_user_by_id(class_obj.teacher_id) if class_obj.teacher_id else None
                    teacher_name = teacher.full_name if teacher else ""
                    
                    # Get student count
                    students_in_class = db.get_students_by_class(class_obj.id)
                    student_count = len(students_in_class)
                    
                    # Apply filters
                    if class_name_search and class_name_search.lower() not in class_obj.name.lower():
                        continue
                    if teacher_search and teacher_search.lower() not in teacher_name.lower():
                        continue
                    if academic_year_search and academic_year_search not in class_obj.academic_year:
                        continue
                    if student_count < min_students:
                        continue
                        
                    filtered_classes.append(class_obj)
                
                if filtered_classes:
                    st.success(f"🔎 Tìm thấy {len(filtered_classes)} lớp học phù hợp")
                    
                    # Export search results
                    if st.button("📊 Xuất danh sách lớp (Excel)", key="export_class_search"):
                        export_class_search_results(db, filtered_classes)
                    
                    st.session_state.filtered_classes = filtered_classes
                    classes = filtered_classes  # Use filtered classes for display
                else:
                    st.info("Không tìm thấy lớp học nào phù hợp với tiêu chí tìm kiếm")
                    return

    # Display class information
    for class_info in classes:
        with st.expander(f"🏫 {class_info.name} - Năm học {class_info.academic_year}", expanded=True):
            tabs = st.tabs([
                get_text("pages.classes.general_info", "Thông tin chung"), 
                get_text("pages.classes.student_list", "Danh sách học sinh"), 
                get_text("pages.classes.academic_history", "Lịch sử học tập"),
                "📝 Ghi chú học sinh"
            ])

            # Tab Thông tin chung
            with tabs[0]:
                # Teacher information
                teacher = db.get_user_by_id(class_info.teacher_id)

                # Edit class button (only for admin and teachers)
                if user_role in ['admin', 'teacher']:
                    col1, col2 = st.columns([3, 1])
                    with col2:
                        if st.button(get_text("pages.classes.edit_class", "✏️ Chỉnh sửa"), key=f"edit_{class_info.id}"):
                            st.session_state[f'editing_class_{class_info.id}'] = True

                # Edit class form
                if user_role in ['admin', 'teacher'] and st.session_state.get(f'editing_class_{class_info.id}', False):
                    with st.form(f"edit_class_{class_info.id}"):
                        st.subheader("Chỉnh sửa thông tin lớp")

                        # Get list of teachers
                        teachers = db.get_teachers()
                        teacher_options = [f"{t.full_name} (ID: {t.id})" for t in teachers]
                        current_teacher = f"{teacher.full_name} (ID: {teacher.id})"
                        
                        # Find current teacher index safely
                        try:
                            current_teacher_index = teacher_options.index(current_teacher)
                        except ValueError:
                            # If current teacher not found, default to first option
                            current_teacher_index = 0
                            if teacher_options:
                                st.warning(f"Giáo viên hiện tại '{current_teacher}' không có trong danh sách. Đã chọn giáo viên đầu tiên.")

                        new_name = st.text_input("Tên lớp", value=class_info.name, key=f"edit_name_{class_info.id}")
                        new_teacher = st.selectbox(
                            "Giáo viên chủ nhiệm",
                            options=teacher_options,
                            index=current_teacher_index,
                            key=f"edit_teacher_{class_info.id}"
                        )
                        # Academic year dropdown for editing
                        year_options = [f"{year}-{year+1}" for year in range(2000, 2041)]
                        current_year_index = 0
                        if class_info.academic_year in year_options:
                            current_year_index = year_options.index(class_info.academic_year)
                        new_academic_year = st.selectbox(
                            "Năm học", 
                            options=year_options,
                            index=current_year_index,
                            key=f"edit_year_{class_info.id}"
                        )
                        new_notes = st.text_area("Ghi chú", value=class_info.notes, key=f"edit_notes_{class_info.id}")

                        if st.form_submit_button("Lưu thay đổi"):
                            try:
                                teacher_id = int(new_teacher.split("ID: ")[1].rstrip(")"))
                                class_data = {
                                    "name": new_name,
                                    "teacher_id": teacher_id,
                                    "academic_year": new_academic_year,
                                    "notes": new_notes
                                }
                                if db.update_class(class_info.id, class_data):
                                    show_success("Cập nhật thông tin lớp thành công!")
                                    st.session_state[f'editing_class_{class_info.id}'] = False
                                    st.rerun()
                                else:
                                    show_error("Không thể cập nhật thông tin lớp")
                            except Exception as e:
                                show_error(f"Lỗi: {str(e)}")

                # Display current teacher info
                st.subheader("👨‍🏫 Giáo viên chủ nhiệm")
                st.write(f"Họ và tên: {teacher.full_name}")
                st.write(f"Email: {teacher.email if teacher.email else 'Chưa cập nhật'}")

            # Tab Danh sách học sinh
            with tabs[1]:
                students = db.get_students_by_class(class_info.id)

                # Add student management controls
                if user_role in ['admin', 'teacher']:
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        if students:
                            st.info(f"👥 Số lượng học sinh: {len(students)} học sinh")
                        else:
                            st.info("Chưa có học sinh trong lớp")
                    with col2:
                        if st.button("➕ Thêm học sinh", key=f"add_student_{class_info.id}"):
                            st.session_state[f'adding_student_{class_info.id}'] = True

                # Form to add student to class
                if user_role in ['admin', 'teacher'] and st.session_state.get(f'adding_student_{class_info.id}', False):
                    with st.form(f"add_student_form_{class_info.id}"):
                        st.subheader("Thêm học sinh vào lớp")
                        
                        # Get list of unassigned students
                        unassigned_students = db.get_unassigned_students()
                        
                        if unassigned_students:
                            student_options = [f"{s.full_name} (ID: {s.id})" for s in unassigned_students]
                            
                            selected_student = st.selectbox(
                                "Chọn học sinh chưa có lớp:",
                                options=student_options,
                                key=f"student_select_{class_info.id}"
                            )
                            
                            reason = st.text_input(
                                "Lý do thêm vào lớp:",
                                value="Thêm mới vào lớp",
                                key=f"reason_{class_info.id}"
                            )

                            col1, col2 = st.columns(2)
                            with col1:
                                if st.form_submit_button("✅ Thêm vào lớp"):
                                    try:
                                        student_id = int(selected_student.split("ID: ")[1].rstrip(")"))
                                        if db.add_student_to_class(student_id, class_info.id, reason):
                                            show_success("Thêm học sinh vào lớp thành công!")
                                            st.session_state[f'adding_student_{class_info.id}'] = False
                                            st.rerun()
                                        else:
                                            show_error("Không thể thêm học sinh vào lớp")
                                    except Exception as e:
                                        show_error(f"Lỗi: {str(e)}")
                            with col2:
                                if st.form_submit_button("❌ Hủy"):
                                    st.session_state[f'adding_student_{class_info.id}'] = False
                                    st.rerun()
                        else:
                            st.info("Không có học sinh chưa được phân lớp")
                            if st.form_submit_button("❌ Đóng"):
                                st.session_state[f'adding_student_{class_info.id}'] = False
                                st.rerun()

                # Display student list in a table format
                if students:
                    student_data = []
                    for student in students:
                        # Only include student in list if user has access
                        if check_student_access(student.id):
                            student_data.append({
                                "Họ và tên": student.full_name,
                                "Nhà T": student.nha_chu_t_info if student.nha_chu_t_info else "Chưa cập nhật",
                                "Sức khỏe khi vào làng": getattr(student, 'health_on_admission', '') or 'Chưa cập nhật',
                                "Đặc điểm sơ bộ": getattr(student, 'initial_characteristics', '') or 'Chưa cập nhật',
                                "Email": student.email,
                            })

                    if student_data:
                        # Display student table
                        st.table(pd.DataFrame(student_data))

                        # Add remove student buttons
                        if user_role in ['admin', 'teacher']:
                            st.write("##### Quản lý học sinh trong lớp")
                            for student in students:
                                col1, col2, col3 = st.columns([2, 1, 1])
                                with col1:
                                    st.write(f"👨‍🎓 {student.full_name}")
                                with col2:
                                    if st.button("📋 Lịch sử", key=f"history_{student.id}"):
                                        history = db.get_student_class_history(student.id)
                                        if history:
                                            st.write(f"**Lịch sử lớp học của {student.full_name}:**")
                                            for record in history:
                                                st.write(f"- {record['class_name']} ({record['academic_year']}) - {record['start_date']} đến {record['end_date']}")
                                        else:
                                            st.info("Chưa có lịch sử lớp học")
                                with col3:
                                    if st.button("🚫 Rời lớp", key=f"remove_{student.id}"):
                                        if db.remove_student_from_class(student.id, "Chuyển ra khỏi lớp"):
                                            show_success(f"Đã chuyển {student.full_name} ra khỏi lớp!")
                                            st.rerun()
                                        else:
                                            show_error("Không thể chuyển học sinh ra khỏi lớp")
                    else:
                        st.info("Không có quyền xem thông tin học sinh trong lớp này")

            # Tab Lịch sử học tập
            with tabs[2]:
                if user_role in ['admin', 'teacher']:
                    st.markdown("### 📚 Lịch sử học tập")

                    # Search filters
                    st.markdown("#### 🔍 Bộ lọc tìm kiếm")
                    col1, col2 = st.columns(2)
                    with col1:
                        search_name = st.text_input("Tên học sinh", key=f"search_student_name_history_class_{class_info.id}")
                    with col2:
                        use_date_filter = st.checkbox("Lọc theo thời gian", key=f"date_filter_history_class_{class_info.id}")
                        if use_date_filter:
                            date_range = st.date_input(
                                "Khoảng thời gian",
                                value=(datetime.now() - timedelta(days=365), datetime.now()),
                                key=f"date_range_history_class_{class_info.id}"
                            )

                    students = db.get_students_by_class(class_info.id)

                    # Search history
                    history_filters = {}
                    if search_name:
                        history_filters['student_name'] = search_name
                    if use_date_filter and len(date_range) == 2:
                        history_filters['from_date'] = date_range[0].isoformat()
                        history_filters['to_date'] = date_range[1].isoformat()

                    # Display history for each student
                    for student in students:
                        if check_student_access(student.id):
                            st.markdown(f"#### 📋 {student.full_name}")
                            history = db.search_student_class_history(
                                student_id=student.id,
                                **history_filters
                            )
                            if history:
                                history_data = pd.DataFrame(history)
                                st.table(history_data)
                            else:
                                st.info("Không tìm thấy lịch sử lớp học phù hợp")

            # Tab Ghi chú học sinh
            with tabs[3]:
                if user_role in ['admin', 'teacher']:
                    st.markdown("### 📝 Ghi chú học sinh")
                    students = db.get_students_by_class(class_info.id)
                    
                    if students:
                        # Select student to add note
                        student_options = [f"{s.full_name} (ID: {s.id})" for s in students]
                        selected_student_option = st.selectbox(
                            "Chọn học sinh để ghi chú:",
                            options=student_options,
                            key=f"note_student_select_{class_info.id}"
                        )
                        
                        if selected_student_option:
                            student_id = int(selected_student_option.split("ID: ")[1].rstrip(")"))
                            selected_student = next((s for s in students if s.id == student_id), None)
                            
                            if selected_student:
                                st.write(f"**Ghi chú cho học sinh: {selected_student.full_name}**")
                                
                                # Form to add new note
                                with st.form(f"add_note_form_{class_info.id}_{student_id}"):
                                    note_content = st.text_area(
                                        "Nội dung ghi chú:",
                                        placeholder="Nhập ghi chú về học sinh...",
                                        height=100,
                                        key=f"note_content_{class_info.id}_{student_id}"
                                    )
                                    
                                    col1, col2 = st.columns([2, 1])
                                    with col1:
                                        note_type = st.selectbox(
                                            "Loại ghi chú:",
                                            options=["Học tập", "Hành vi", "Sức khỏe", "Gia đình", "Khác"],
                                            key=f"note_type_{class_info.id}_{student_id}"
                                        )
                                    with col2:
                                        is_important = st.checkbox(
                                            "Quan trọng",
                                            key=f"note_important_{class_info.id}_{student_id}"
                                        )
                                    
                                    if st.form_submit_button("💾 Lưu ghi chú"):
                                        if note_content.strip():
                                            # Add note to database
                                            note_data = {
                                                'student_id': student_id,
                                                'teacher_id': st.session_state.user.id,
                                                'class_id': class_info.id,
                                                'content': note_content.strip(),
                                                'note_type': note_type,
                                                'is_important': is_important,
                                                'created_at': datetime.now().isoformat()
                                            }
                                            
                                            if db.add_student_note(note_data):
                                                show_success("Đã lưu ghi chú thành công!")
                                                st.rerun()
                                            else:
                                                show_error("Không thể lưu ghi chú")
                                        else:
                                            show_error("Vui lòng nhập nội dung ghi chú")
                                
                                # Display existing notes for this student
                                st.markdown("#### 📋 Ghi chú hiện có")
                                existing_notes = db.get_student_notes(student_id, class_info.id)
                                
                                if existing_notes:
                                    for note in existing_notes:
                                        # Create expander for each note
                                        importance_icon = "⚠️" if note.get('is_important') else "📝"
                                        note_date = note.get('created_at', '')[:10] if note.get('created_at') else ''
                                        
                                        with st.expander(f"{importance_icon} {note.get('note_type', 'Khác')} - {note_date}"):
                                            st.write(f"**Giáo viên:** {note.get('teacher_name', 'Không rõ')}")
                                            st.write(f"**Ngày tạo:** {note_date}")
                                            st.write(f"**Loại:** {note.get('note_type', 'Khác')}")
                                            if note.get('is_important'):
                                                st.warning("⚠️ **Ghi chú quan trọng**")
                                            st.write(f"**Nội dung:** {note.get('content', '')}")
                                            
                                            # Delete note option for admin/teacher who created it
                                            if user_role == 'admin' or note.get('teacher_id') == st.session_state.user.id:
                                                if st.button(f"🗑️ Xóa ghi chú", key=f"delete_note_{note.get('id')}"):
                                                    if db.delete_student_note(note.get('id')):
                                                        show_success("Đã xóa ghi chú!")
                                                        st.rerun()
                                                    else:
                                                        show_error("Không thể xóa ghi chú")
                                else:
                                    st.info("Chưa có ghi chú nào cho học sinh này")
                    else:
                        st.info("Không có học sinh trong lớp")
                else:
                    st.warning("Chỉ giáo viên và quản trị viên mới có thể xem ghi chú học sinh")

def export_class_search_results(db, classes):
    """Export class search results to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách lớp học"
    
    # Add title
    ws.merge_cells('A1:F1')
    ws['A1'] = "DANH SÁCH LỚP HỌC"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    headers = ['TT', 'Tên lớp', 'Giáo viên chủ nhiệm', 'Năm học', 'Số học sinh', 'Ghi chú']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Format header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    for i, class_obj in enumerate(classes, 1):
        teacher = db.get_user_by_id(class_obj.teacher_id) if class_obj.teacher_id else None
        teacher_name = teacher.full_name if teacher else "Chưa phân công"
        students_count = len(db.get_students_by_class(class_obj.id))
        
        ws.append([
            i,
            class_obj.name,
            teacher_name,
            class_obj.academic_year,
            students_count,
            class_obj.notes or ""
        ])
    
    # Set column widths
    column_widths = [5, 25, 25, 15, 15, 30]
    for i, width in enumerate(column_widths, 1):
        column_letter = chr(64 + i)
        ws.column_dimensions[column_letter].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Danh_sach_lop_hoc_{timestamp}.xlsx"
    
    st.download_button(
        label="📥 Tải xuống danh sách lớp (Excel)",
        data=output.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_class_list"
    )
    st.success("Đã tạo file Excel thành công!")

if __name__ == "__main__":
    render()